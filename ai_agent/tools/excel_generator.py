from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

class ExcelGenerator:
    """Генератор Excel документов для различных типов отчётов"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _style_header(self, ws, headers):
        """Стилизировать заголовки"""
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.border
    
    def _style_rows(self, ws, start_row, end_row, num_cols):
        """Стилизировать строки данных"""
        for row_num in range(start_row, end_row + 1):
            for col_num in range(1, num_cols + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = self.border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    def create_product_analysis(self, products_data: list) -> str:
        """Создать отчёт анализа товаров с маркетплейса
        
        products_data: list[{
            'name': str,
            'marketplace': str,
            'price': float,
            'rating': float,
            'reviews': int,
            'available': bool,
            'description': str
        }]
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Анализ товаров"
        
        headers = ["№", "Название товара", "Маркетплейс", "Цена (₽)", "Рейтинг", "Отзывы", "Статус"]
        self._style_header(ws, headers)
        
        for idx, product in enumerate(products_data, 1):
            row = idx + 1
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = product.get('name', '')
            ws.cell(row=row, column=3).value = product.get('marketplace', '')
            ws.cell(row=row, column=4).value = product.get('price', 0)
            ws.cell(row=row, column=5).value = product.get('rating', 0)
            ws.cell(row=row, column=6).value = product.get('reviews', 0)
            ws.cell(row=row, column=7).value = "Доступно" if product.get('available') else "Нет в наличии"
        
        self._style_rows(ws, 2, len(products_data) + 1, len(headers))
        
        # Установить ширину колонок
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 15
        
        filename = f"product_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join("files", filename)
        os.makedirs("files", exist_ok=True)
        wb.save(filepath)
        
        return filepath
    
    def create_price_comparison(self, comparison_data: list) -> str:
        """Создать отчёт сравнения цен
        
        comparison_data: list[{
            'name': str,
            'ozon': float,
            'wildberries': float,
            'yandex_market': float,
            'min_price': float,
            'best_marketplace': str
        }]
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Сравнение цен"
        
        headers = ["Товар", "Озон (₽)", "Валберис (₽)", "Яндекс Маркет (₽)", "Минимальная цена", "Лучшее предложение"]
        self._style_header(ws, headers)
        
        for idx, item in enumerate(comparison_data, 1):
            row = idx + 1
            ws.cell(row=row, column=1).value = item.get('name', '')
            ws.cell(row=row, column=2).value = item.get('ozon', '-')
            ws.cell(row=row, column=3).value = item.get('wildberries', '-')
            ws.cell(row=row, column=4).value = item.get('yandex_market', '-')
            ws.cell(row=row, column=5).value = item.get('min_price', '-')
            ws.cell(row=row, column=6).value = item.get('best_marketplace', '-')
        
        self._style_rows(ws, 2, len(comparison_data) + 1, len(headers))
        
        ws.column_dimensions['A'].width = 35
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15
        
        filename = f"price_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join("files", filename)
        os.makedirs("files", exist_ok=True)
        wb.save(filepath)
        
        return filepath
    
    def create_custom_report(self, title: str, columns: list, data: list) -> str:
        """Создать пользовательский отчёт
        
        columns: list[str] - названия колонок
        data: list[list] - данные для каждой строки
        """
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel ограничивает имя листа на 31 символ
        
        self._style_header(ws, columns)
        
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx + 1, column=col_idx).value = value
        
        self._style_rows(ws, 2, len(data) + 1, len(columns))
        
        # Авто-ширина колонок
        for col_num, col_title in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        
        filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join("files", filename)
        os.makedirs("files", exist_ok=True)
        wb.save(filepath)
        
        return filepath
    
    def create_debt_report(self, debts_data: list) -> str:
        """Создать отчёт о задолженности
        
        debts_data: list[{
            'client': str,
            'amount': float,
            'due_date': str,
            'days_overdue': int,
            'status': str
        }]
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Задолженность"
        
        headers = ["Клиент", "Сумма (₽)", "Срок оплаты", "Дней просрочки", "Статус"]
        self._style_header(ws, headers)
        
        total_debt = 0
        overdue_debt = 0
        
        for idx, debt in enumerate(debts_data, 1):
            row = idx + 1
            ws.cell(row=row, column=1).value = debt.get('client', '')
            ws.cell(row=row, column=2).value = debt.get('amount', 0)
            ws.cell(row=row, column=3).value = debt.get('due_date', '')
            ws.cell(row=row, column=4).value = debt.get('days_overdue', 0)
            ws.cell(row=row, column=5).value = debt.get('status', '')
            
            total_debt += debt.get('amount', 0)
            if debt.get('days_overdue', 0) > 0:
                overdue_debt += debt.get('amount', 0)
        
        self._style_rows(ws, 2, len(debts_data) + 1, len(headers))
        
        # Добавить итоговые строки
        summary_row = len(debts_data) + 3
        ws.cell(row=summary_row, column=1).value = "ИТОГО:"
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        ws.cell(row=summary_row, column=2).value = total_debt
        ws.cell(row=summary_row, column=2).font = Font(bold=True)
        
        ws.cell(row=summary_row + 1, column=1).value = "Просроченная задолженность:"
        ws.cell(row=summary_row + 1, column=1).font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=2).value = overdue_debt
        ws.cell(row=summary_row + 1, column=2).font = Font(bold=True, color="FF0000")
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        filename = f"debt_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join("files", filename)
        os.makedirs("files", exist_ok=True)
        wb.save(filepath)
        
        return filepath
